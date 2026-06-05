import openpyxl
import os
from pathlib import Path
from src.common_library import helper_functions as hf

def read_spreadsheet(root_dir, sub_folder, file_name, sheet_name):
    """
    This function reads the data from the spreadsheet into a dictionary.

    Args:
        root_dir (string): root directory for the project
        sub_folder(string): the folder containing the spreadsheet to process.
        file_name (string): the name of the file
        sheet_name (string): name of the sheet in the workbook to read.

    Returns:
        results (list of tuples): Each tuple contains 2 elements: [0] song title, and [1] artist.
    """

    # TODO: give the user the ability to specify an environment when there are multiple environments
    results = []
    file_path = hf.build_file_name(root_dir, sub_folder, file_name)
    # file_path = os.environ.get("steps_spreadsheet_directory_test")
    # xlsx_file = Path(file_path, 'steps.xlsx')
    xlsx_file = Path(file_path)

    try:
        wb_obj = openpyxl.load_workbook(xlsx_file)
    except FileNotFoundError as error:
        error_message = " ".join(["File not found."])
        print(error_message)
        raise error

    # get the worksheet
    try:
        mySheet = wb_obj[sheet_name]
    except KeyError as error:
        error_message = " ".join(["Worksheet not found."])
        print(error_message)
        raise error

    try:
        col_names = []
        for column in mySheet.iter_cols(1, mySheet.max_column):
            col_names.append(column[0].value)

        #### determine indexes for Date and Steps columns
        for i in range(0, mySheet.max_column):
            if col_names[i] == "Title": title_idx = i
            if col_names[i] == "Artist": artist_idx = i

        # assumption: first row contains column headers
        # actual data starts in row 2.
        for row in mySheet.iter_rows(min_row=2):
            results.append((row[title_idx].value, row[artist_idx].value))

    except Exception as error:
        raise error

    return results
